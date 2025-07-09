from django.shortcuts import render
from django.shortcuts import render, redirect
from .forms import *
from .models import StudentAdmission, StudentAcademicRecord, Class, AcademicYear
from core.models import UserProfile
from django.contrib.auth.decorators import login_required
from django.utils.timezone import now
from django.http import FileResponse
import os
from django.conf import settings
from django.shortcuts import get_object_or_404 
from django.http import HttpResponseForbidden
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from openpyxl import load_workbook
from django.utils.dateparse import parse_date
from dateutil.parser import parse as smart_date_parse
from datetime import datetime, date
from django.http import HttpResponse
from docxtpl import DocxTemplate

@login_required
def admission_home(request):
    return render(request, 'admission/admission_home.html')

@login_required
def admit_student(request):
    school = request.user.userprofile.school

    # Generate next admission number
    last_student = StudentAdmission.objects.filter(school=school).order_by('-id').first()
    if last_student and last_student.admission_no.isdigit():
        last_adm_no = int(last_student.admission_no)
        next_adm_no = str(last_adm_no + 1)
    else:
        next_adm_no = '1001'

    if request.method == 'POST':
        form = StudentAdmissionForm(request.POST, request.FILES)
        if form.is_valid():
            student = form.save(commit=False)
            student.school = school
            student.admission_no = next_adm_no
            student.is_active = True

            # ✅ Format and clean inputs
            student.full_name = str(student.full_name).strip().title()
            student.gender = str(student.gender).strip().capitalize()
            student.category = str(student.category).strip().upper()
            student.religion = str(student.religion).strip().title()
            student.mobile_no = str(student.mobile_no).strip()
            student.whatsapp_no = str(student.whatsapp_no).strip()
            student.aadhar_no = str(student.aadhar_no).strip()
            student.father_name = str(student.father_name).strip().title() if student.father_name else ""
            student.mother_name = str(student.mother_name).strip().title() if student.mother_name else ""
            student.father_profession = str(student.father_profession).strip().title() if student.father_profession else ""

            student.save()

            # ✅ Academic record creation
            academic_year = form.cleaned_data['academic_year']
            class_enrolled = form.cleaned_data['class_enrolled']
            section = str(form.cleaned_data.get('section', '')).strip().upper()

            StudentAcademicRecord.objects.create(
                student=student,
                academic_year=academic_year,
                class_enrolled=class_enrolled,
                section=section,
                school=school
            )

            return redirect('student_list')
    else:
        form = StudentAdmissionForm(initial={'admission_no': next_adm_no})

    return render(request, 'admission/admit_student.html', {'form': form})

@login_required
def student_list(request):
    school = request.user.userprofile.school
    if request.user.is_superuser:
        students = StudentAdmission.objects.filter(is_active=True)
        # print("Found students:", students) 
    else:
        students = StudentAdmission.objects.filter(
        school=school,
        is_active=True
        )
    
    # Filters
    class_id = request.GET.get('class_id')
    year_id = request.GET.get('year_id')
    name_query = request.GET.get('name',"")

    academic_records = StudentAcademicRecord.objects.filter(school=school)

    if class_id:
        academic_records = academic_records.filter(class_enrolled_id=class_id)

    if year_id:
        academic_records = academic_records.filter(academic_year_id=year_id)

    if name_query:
        students = students.filter(full_name__icontains=name_query)

    # Match academic records with students
    student_ids = academic_records.values_list('student_id', flat=True)
    students = students.filter(id__in=student_ids)

    classes = Class.objects.filter(school=school, is_active=True)
    years = AcademicYear.objects.filter(school=school)

    return render(request, 'admission/student_list.html', {
        'students': students,
        'classes': classes,
        'years': years,
        'selected_class': class_id,
        'selected_year': year_id,
        'name_query': name_query
    })

@login_required
def add_class(request):
    form = ClassForm(request.POST or None)
    if form.is_valid():
        cls = form.save(commit=False)
        cls.school = request.user.userprofile.school
        cls.save()
        return redirect('admission_home')
    return render(request, 'admission/add_class.html', {'form': form})

@login_required
def add_academic_year(request):
    form = AcademicYearForm(request.POST or None)
    if form.is_valid():
        year = form.save(commit=False)
        year.school = request.user.userprofile.school
        year.save()
        return redirect('admission_home')
    return render(request, 'admission/add_academic_year.html', {'form': form})

@login_required
def class_list(request):
    school = request.user.userprofile.school
    classes = Class.objects.filter(school=school)
    return render(request, 'admission/class_list.html', {'classes': classes})

@login_required
def academic_year_list(request):
    school = request.user.userprofile.school
    years = AcademicYear.objects.filter(school=school)
    return render(request, 'admission/academic_year_list.html', {'years': years})

@login_required
def edit_class(request, class_id):
    school = request.user.userprofile.school
    cls = Class.objects.get(id=class_id, school=school)
    form = ClassForm(request.POST or None, instance=cls)
    if form.is_valid():
        form.save()
        return redirect('class_list')
    return render(request, 'admission/edit_class.html', {'form': form})

@login_required
def delete_class(request, class_id):
    school = request.user.userprofile.school
    cls = Class.objects.get(id=class_id, school=school)
    if request.method == 'POST':
        cls.delete()
        return redirect('class_list')
    return render(request, 'admission/delete_confirm.html', {'object': cls, 'type': 'Class'})

@login_required
def edit_year(request, year_id):
    school = request.user.userprofile.school
    year = AcademicYear.objects.get(id=year_id, school=school)
    form = AcademicYearForm(request.POST or None, instance=year)
    if form.is_valid():
        form.save()
        return redirect('academic_year_list')
    return render(request, 'admission/edit_year.html', {'form': form})

@login_required
def delete_year(request, year_id):
    school = request.user.userprofile.school
    year = AcademicYear.objects.get(id=year_id, school=school)
    if request.method == 'POST':
        year.delete()
        return redirect('academic_year_list')
    return render(request, 'admission/delete_confirm.html', {'object': year, 'type': 'Academic Year'})

@login_required
def edit_student_academic_record(request, student_id):
    """
    View to edit the latest academic record of a student.
    Only used for admin corrections.
    """
    student = get_object_or_404(StudentAdmission, id=student_id)
    academic_record = student.academic_records.last()

    if not academic_record:
        return render(request, "admission/no_record_found.html", {"student": student})

    form = StudentAcademicRecordForm(request.POST or None, instance=academic_record)

    if request.method == 'POST' and form.is_valid():
        updated_record = form.save(commit=False)
        updated_record.student = student
        updated_record.school = academic_record.school  # keep original school
        updated_record.save()
        return redirect('student_list')

    return render(request, 'admission/edit_academic_record.html', {
        'form': form,
        'student': student,
        'academic_record': academic_record,
    })

@login_required
def view_academic_records(request, student_id):
    student = get_object_or_404(StudentAdmission, id=student_id)
    records = student.academic_records.select_related('academic_year', 'class_enrolled', 'promoted_to')

    return render(request, 'admission/view_academic_records.html', {
        'student': student,
        'records': records
    })

@login_required
def student_profile(request, student_id):
    student = get_object_or_404(StudentAdmission, id=student_id)
    records = student.academic_records.select_related('academic_year', 'class_enrolled', 'promoted_to')
    
    return render(request, 'admission/student_profile.html', {
        'student': student,
        'records': records
    })

@login_required
def edit_student(request, student_id):
    student = get_object_or_404(StudentAdmission, id=student_id)
    academic_record = student.academic_records.first()  # adjust if multiple years

    if request.method == 'POST':
        form = StudentEditForm(request.POST, request.FILES, instance=student, academic_record=academic_record)
        if form.is_valid():
            form.save()
            # update academic record too
            if academic_record:
                academic_record.academic_year = form.cleaned_data['academic_year']
                academic_record.class_enrolled = form.cleaned_data['class_enrolled']
                academic_record.save()

            messages.success(request, 'Student and Academic info updated successfully.')
            return redirect('student_profile', student_id=student.id)
    else:
        form = StudentEditForm(instance=student, academic_record=academic_record)

    return render(request, 'admission/edit_student.html', {
        'form': form,
        'student': student
    })

@login_required
def soft_delete_student(request, student_id):
    student = get_object_or_404(StudentAdmission, id=student_id)
    if request.user.userprofile.school != student.school:
        return HttpResponseForbidden("You don't have permission to delete this student.")
    
    if request.method == 'POST':
        student.is_active = False
        student.save()
        return redirect('student_list')

@login_required
def import_students_excel(request):
    if request.method == 'POST' and request.FILES.get('student_file'):
        excel_file = request.FILES['student_file']
        wb = load_workbook(excel_file)
        ws = wb.active
        school = request.user.userprofile.school
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        for row in rows:
            try:
                (
                    full_name, date_of_birth, gender, category,
                    religion, mobile_no, whatsapp_no, aadhar_no,
                    admission_date, father_name, mother_name, father_profession,
                    academic_year_str, class_name, section
                ) = row

                # ✅ Normalize string fields
                full_name = str(full_name).strip().title() if full_name else ""
                gender = str(gender).strip().capitalize() if gender else ""
                category = str(category).strip().upper() if category else ""
                religion = str(religion).strip().title() if religion else ""
                mobile_no = str(mobile_no).strip() if mobile_no else ""
                whatsapp_no = str(whatsapp_no).strip() if whatsapp_no else ""
                aadhar_no = str(aadhar_no).strip() if aadhar_no else ""
                father_name = str(father_name).strip().title() if father_name else ""
                mother_name = str(mother_name).strip().title() if mother_name else ""
                father_profession = str(father_profession).strip().title() if father_profession else ""
                section = str(section).strip().upper() if section else ""
                academic_year_str = str(academic_year_str).strip()
                class_name = str(class_name).strip()

                # ✅ Parse DOB
                if isinstance(date_of_birth, (datetime, date)):
                    dob = date_of_birth
                else:
                    try:
                        dob = smart_date_parse(str(date_of_birth)).date()
                    except:
                        messages.error(request, f"{full_name}: Invalid or unreadable Date of Birth '{date_of_birth}'.")
                        continue

                # ✅ Parse Admission Date
                if isinstance(admission_date, (datetime, date)):
                    adm_date = admission_date
                else:
                    try:
                        adm_date = smart_date_parse(str(admission_date)).date()
                    except:
                        messages.error(request, f"{full_name}: Invalid or unreadable Admission Date '{admission_date}'.")
                        continue

                # ✅ Validate Gender
                if gender not in ["Male", "Female", "Other"]:
                    messages.error(request, f"{full_name}: Invalid gender '{gender}'.")
                    continue

                # ✅ Validate Category
                if category not in ["GEN", "OBC", "SC", "ST"]:
                    messages.error(request, f"{full_name}: Invalid category '{category}'.")
                    continue

                # ✅ Normalize class name (handle "6" as "Class 6")
                normalized_class_name = class_name if class_name.lower().startswith("class") else f"Class {class_name}"

                # ✅ Lookup academic year and class
                academic_year = AcademicYear.objects.get(name=academic_year_str, school=school)
                class_enrolled = Class.objects.get(name=normalized_class_name, school=school)

                # ✅ Generate admission number
                last_student = StudentAdmission.objects.filter(school=school).order_by('-id').first()
                next_adm_no = str(int(last_student.admission_no) + 1) if last_student and last_student.admission_no.isdigit() else '1001'

                # ✅ Create Student
                student = StudentAdmission.objects.create(
                    full_name=full_name,
                    date_of_birth=dob,
                    gender=gender,
                    category=category,
                    religion=religion,
                    mobile_no=mobile_no,
                    whatsapp_no=whatsapp_no,
                    aadhar_no=aadhar_no,
                    admission_date=adm_date,
                    father_name=father_name,
                    mother_name=mother_name,
                    father_profession=father_profession,
                    school=school,
                    is_active=True,
                    admission_no=next_adm_no
                )

                # ✅ Create Academic Record
                StudentAcademicRecord.objects.create(
                    student=student,
                    academic_year=academic_year,
                    class_enrolled=class_enrolled,
                    section=section,
                    school=school
                )

                messages.success(request, f"{full_name} imported successfully.")

            except AcademicYear.DoesNotExist:
                messages.error(request, f"{full_name}: Academic Year '{academic_year_str}' not found.")
            except Class.DoesNotExist:
                messages.error(request, f"{full_name}: Class '{class_name}' not found.")
            except Exception as e:
                messages.error(request, f"{full_name}: Error importing - {str(e)}")

        return redirect('import_students_excel')

    return render(request, 'admission/import_students_excel.html')

@login_required
def download_excel_template(request):
    file_path = os.path.join(settings.BASE_DIR, 'static/admission/student_import_template.xlsx')
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename='student_template.xlsx')

@login_required
def generate_admission_form(request, student_id):
    student = StudentAdmission.objects.get(id=student_id)
    academic_record = StudentAcademicRecord.objects.filter(student=student).order_by('-academic_year__start_date').first()
    
    template_path = os.path.join('static', 'admission', 'admission_template.docx')
    doc = DocxTemplate(template_path)

    context = {
        'full_name': student.full_name,
        'ssr_no': student.ssr_no,
        'academic_year': academic_record.academic_year.name if academic_record else '',
        'gender': student.gender,
        'date_of_birth': student.date_of_birth.strftime('%d %B %Y') if student.date_of_birth else '',
        'admission_date': student.admission_date.strftime('%d %B %Y') if student.admission_date else '',
        'admission_no': student.admission_no,
        'father_name': student.father_name,
        'mother_name': student.mother_name, 
        'father_profession': student.father_profession,
        'aadhar_no': student.aadhar_no,
        'mobile_no': student.mobile_no,
        'whatsapp_no': student.whatsapp_no,
        'religion': student.religion,
        'category': student.category,
        'address': student.address if student.address else '',
        'class_enrolled': academic_record.class_enrolled.name if academic_record else '',
        'section': academic_record.section if academic_record and academic_record.section else '',
    }

    doc.render(context)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    response['Content-Disposition'] = f'attachment; filename=admission_form_{student.admission_no}.docx'
    doc.save(response)
    return response


