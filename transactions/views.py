# transactions/views.py
from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import *
from .models import *
from django.contrib.auth.decorators import login_required
from admission.models import School 
from django.db.models import Q
from django.utils.dateparse import parse_date
import openpyxl
from django.http import HttpResponse
from decimal import Decimal
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from django.forms import modelformset_factory

@login_required
def transactions_home(request):
    return render(request, 'transactions/transactions_home.html')

@login_required
def add_manual_transaction(request):
     # adjust import if needed

    user_school = request.user.userprofile.school
    form = ManualTransactionForm(request.POST or None)
    form.fields['debit_account'].queryset = form.fields['debit_account'].queryset.filter(school=user_school)
    form.fields['credit_account'].queryset = form.fields['credit_account'].queryset.filter(school=user_school)

    if request.method == 'POST' and form.is_valid():
        cd = form.cleaned_data

        if cd['debit_account'] == cd['credit_account']:
            messages.error(request, "Debit and Credit accounts cannot be the same.")
            return render(request, 'transactions/add_manual_transaction.html', {'form': form})

        Transaction.objects.create(
            date=cd['date'],
            debit_account=cd['debit_account'],
            credit_account=cd['credit_account'],
            amount=cd['amount'],
            remarks=cd['remarks'],
            voucher_type=cd['voucher_type'],
            school=user_school,
            created_by=request.user
        )
        messages.success(request, "Transaction recorded successfully.")
        return redirect('add_manual_transaction')

    return render(request, 'transactions/add_manual_transaction.html', {'form': form})

@login_required
def view_transactions(request):
    user_school = request.user.userprofile.school
    transactions = Transaction.objects.filter(school=user_school).order_by('-date')

    # Filters
    account_id = request.GET.get('account')
    voucher_type = request.GET.get('voucher_type')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if account_id and account_id.isdigit():
        transactions = transactions.filter(
            Q(debit_account_id=account_id) | Q(credit_account_id=account_id)
        )
    if voucher_type:
        transactions = transactions.filter(voucher_type=voucher_type)
    parsed_start = parse_date(start_date) if isinstance(start_date, str) and start_date else None
    if parsed_start:
        transactions = transactions.filter(date__gte=parsed_start)

    parsed_end = parse_date(end_date) if isinstance(end_date, str) and end_date else None
    if parsed_end:
        transactions = transactions.filter(date__lte=parsed_end)


    accounts = AccountHead.objects.filter(school=user_school)

    return render(request, 'transactions/view_transactions.html', {
        'transactions': transactions,
        'accounts': accounts,
        'voucher_type_filter': voucher_type,
        'account_filter': account_id,
        'start_date': start_date,
        'end_date': end_date,
    })


@login_required
def export_transactions_excel(request):
    user_school = request.user.userprofile.school
    transactions = Transaction.objects.filter(school=user_school).order_by('-date')

    # Apply filters (same as in view_transactions)
    account_id = request.GET.get('account')
    voucher_type = request.GET.get('voucher_type')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if account_id and account_id.isdigit():
        transactions = transactions.filter(
            Q(debit_account_id=account_id) | Q(credit_account_id=account_id)
        )
    if voucher_type:
        transactions = transactions.filter(voucher_type=voucher_type)
        
    parsed_start = parse_date(start_date) if isinstance(start_date, str) and start_date else None
    if parsed_start:
        transactions = transactions.filter(date__gte=parsed_start)

    parsed_end = parse_date(end_date) if isinstance(end_date, str) and end_date else None
    if parsed_end:
        transactions = transactions.filter(date__lte=parsed_end)


    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Header
    ws.append(['Date', 'Debit Account', 'Credit Account', 'Amount', 'Voucher Type', 'Remarks'])

    # Data rows
    for txn in transactions:
        ws.append([
            txn.date.strftime('%Y-%m-%d'),
            txn.debit_account.name,
            txn.credit_account.name,
            float(txn.amount),
            txn.get_voucher_type_display(),
            txn.remarks or ''
        ])

    # Set response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="transactions.xlsx"'
    wb.save(response)
    return response

@login_required
def ledger_view(request):
    user_school = request.user.userprofile.school
    accounts = AccountHead.objects.filter(school=user_school)

    selected_account_id = request.GET.get('account')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    selected_account = None
    ledger_entries = []
    running_balance = Decimal('0.00')
    parsed_start = parse_date(start_date) if isinstance(start_date, str) and start_date else None
    parsed_end = parse_date(end_date) if isinstance(end_date, str) and end_date else None


    if selected_account_id and selected_account_id.isdigit():
        selected_account = get_object_or_404(AccountHead, id=selected_account_id, school=user_school)

        transactions = Transaction.objects.filter(
            Q(debit_account=selected_account) | Q(credit_account=selected_account),
            school=user_school
        ).order_by('date', 'id')

        if parsed_start:
            transactions = transactions.filter(date__gte=parsed_start)
        if parsed_end:
            transactions = transactions.filter(date__lte=parsed_end)

        # Initial opening balance
        running_balance = selected_account.opening_balance or Decimal('0.00')

        # Build ledger rows with running balance
        for txn in transactions:
            if txn.debit_account == selected_account:
                debit = txn.amount
                credit = Decimal('0.00')
                running_balance += txn.amount
            else:
                debit = Decimal('0.00')
                credit = txn.amount
                running_balance -= txn.amount

            ledger_entries.append({
                'date': txn.date,
                'debit': debit,
                'credit': credit,
                'remarks': txn.remarks,
                'voucher_type': txn.get_voucher_type_display(),
                'opposite_account': txn.credit_account.name if debit else txn.debit_account.name,
                'running_balance': running_balance
            })

    return render(request, 'transactions/ledger_view.html', {
        'accounts': accounts,
        'selected_account': selected_account,
        'ledger_entries': ledger_entries,
        'start_date': start_date,
        'end_date': end_date
    })

@login_required
def export_ledger_excel(request):
    user_school = request.user.userprofile.school
    account_id = request.GET.get('account')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    parsed_start = parse_date(start_date) if isinstance(start_date, str) and start_date else None
    parsed_end = parse_date(end_date) if isinstance(end_date, str) and end_date else None

    if not account_id or not account_id.isdigit():
        return HttpResponse("Invalid account selected.", status=400)

    account = get_object_or_404(AccountHead, id=account_id, school=user_school)

    transactions = Transaction.objects.filter(
        Q(debit_account=account) | Q(credit_account=account),
        school=user_school
    ).order_by('date', 'id')

    if parsed_start:
        transactions = transactions.filter(date__gte=parsed_start)
    if parsed_end:
        transactions = transactions.filter(date__lte=parsed_end)

    # Prepare workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ledger"

    ws.append([
        'Date', 'Voucher Type', 'Opposite Account',
        'Debit (₹)', 'Credit (₹)', 'Running Balance (₹)', 'Remarks'
    ])

    running_balance = account.opening_balance or Decimal('0.00')

    for txn in transactions:
        if txn.debit_account == account:
            debit = txn.amount
            credit = Decimal('0.00')
            running_balance += debit
            opposite = txn.credit_account.name
        else:
            debit = Decimal('0.00')
            credit = txn.amount
            running_balance -= credit
            opposite = txn.debit_account.name

        ws.append([
            txn.date.strftime('%Y-%m-%d'),
            txn.get_voucher_type_display(),
            opposite,
            float(debit),
            float(credit),
            float(running_balance),
            txn.remarks or ''
        ])

    # Auto column width
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"ledger_{account.name.replace(' ', '_')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def set_opening_balances(request):
    user_school = request.user.userprofile.school

    AccountHeadFormSet = modelformset_factory(
        AccountHead,
        form=AccountHeadBalanceForm,
        extra=0
    )

    queryset = AccountHead.objects.filter(school=user_school).order_by('type', 'name')
    formset = AccountHeadFormSet(request.POST or None, queryset=queryset)

    if request.method == 'POST':
        print("POST received ✅")
        # print("Raw POST data:", request.POST)

        if formset.is_valid():
            print("Formset is valid ✅")
            formset.save()
            messages.success(request, "Opening balances updated successfully.")
            return redirect('set_opening_balances')
        else:
            print("Formset is NOT valid ❌")
            # print("Formset errors:", formset.errors)

    return render(request, 'transactions/set_opening_balances.html', {
        'formset': formset,
    })



@login_required
def create_account_head(request):
    school = request.user.userprofile.school
    form = AccountHeadForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        account_head = form.save(commit=False)
        account_head.school = school
        account_head.is_active = True
        account_head.save()
        return redirect('transactions_home')  # redirect to a listing page or dashboard

    return render(request, 'transactions/create_account_head.html', {'form': form})

@login_required
def list_account_heads(request):
    school = request.user.userprofile.school
    heads = AccountHead.objects.filter(school=school).order_by('type', 'name')
    return render(request, 'transactions/account_head_list.html', {'heads': heads})

# transactions/views.py
# import pandas as pd
# from io import BytesIO
# from django.http import HttpResponse
# from .models import AccountHead

# @login_required
# def export_account_heads_excel(request):
#     school = request.user.userprofile.school
#     heads = AccountHead.objects.filter(school=school)

#     data = []
#     for h in heads:
#         data.append({
#             'Name': h.name,
#             'Type': h.get_type_display(),
#             'Description': h.description,
#             'Opening Balance': float(h.opening_balance),
#             'Status': 'Active' if h.is_active else 'Inactive',
#         })

#     df = pd.DataFrame(data)

#     output = BytesIO()
#     with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
#         df.to_excel(writer, index=False, sheet_name='AccountHeads')
#     output.seek(0)

#     response = HttpResponse(
#         output,
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename=account_heads.xlsx'
#     return response
