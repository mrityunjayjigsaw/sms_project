from django.shortcuts import render
from django.contrib.auth.decorators import login_required

@login_required
def reports_home(request):
    return render(request, 'reports/reports_home.html')


# reports/views.py
from admission.models import *
from django.contrib.auth.decorators import login_required

@login_required
def student_list_report(request):
    user_school = request.user.userprofile.school
    academic_year = request.GET.get('academic_year')
    class_id = request.GET.get('class_id')

    academic_years = AcademicYear.objects.filter(school=user_school)
    classes = Class.objects.filter(school=user_school)

    records = StudentAcademicRecord.objects.filter(school=user_school).select_related('student', 'academic_year', 'class_enrolled')

    if academic_year:
        records = records.filter(academic_year_id=academic_year)
    if class_id:
        records = records.filter(class_enrolled_id=class_id)

    return render(request, 'reports/admission/student_list_report.html', {
        'records': records,
        'academic_years': academic_years,
        'classes': classes,
        'selected_year': academic_year,
        'selected_class': class_id,
    })



# reports/views.py (continued)
import openpyxl
from django.http import HttpResponse
from django.utils.dateparse import parse_date

@login_required
def export_student_list_report(request):
    user_school = request.user.userprofile.school
    academic_year = request.GET.get('academic_year')
    class_id = request.GET.get('class_id')

    records = StudentAcademicRecord.objects.filter(school=user_school).select_related('student', 'academic_year', 'class_enrolled')
    if academic_year:
        records = records.filter(academic_year_id=academic_year)
    if class_id:
        records = records.filter(class_enrolled_id=class_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student List"

    ws.append(['Admission No', 'Full Name', 'Gender', 'Class', 'Academic Year', 'Mobile No'])

    for r in records:
        s = r.student
        ws.append([
            s.admission_no,
            s.full_name,
            s.gender,
            r.class_enrolled.name if r.class_enrolled else '',
            r.academic_year.name if r.academic_year else '',
            s.mobile_no
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="student_list.xlsx"'
    wb.save(response)
    return response

# reports/views.py
from fees.models import StudentFeeDue

from calendar import month_name
import calendar

from django.shortcuts import render
from django.contrib.auth.decorators import login_required

from fees.models import StudentFeeDue, StudentFeePaymentDetail
from django.db.models import Sum
import calendar

@login_required
def fee_defaulter_report(request):
    user_school = request.user.userprofile.school
    academic_year = request.GET.get('academic_year')
    class_id = request.GET.get('class_id')
    month_str = request.GET.get('month')

    academic_years = AcademicYear.objects.filter(school=user_school)
    classes = Class.objects.filter(school=user_school)

    # Safely convert to int if present
    year_id = int(academic_year) if academic_year and academic_year.isdigit() else None
    class_id_val = int(class_id) if class_id and class_id.isdigit() else None

    dues = StudentFeeDue.objects.select_related('student', 'fee_type').filter(student__school=user_school)

    if year_id:
        dues = dues.filter(student__academic_records__academic_year_id=year_id)
    if class_id_val:
        dues = dues.filter(student__academic_records__class_enrolled_id=class_id_val)
    if month_str:
        dues = dues.filter(month__month=int(month_str))

    defaulters = []
    for due in dues:
        paid = StudentFeePaymentDetail.objects.filter(
            payment__student=due.student,
            payment__month=due.month,
            fee_type=due.fee_type
        ).aggregate(total=Sum('amount_paid'))['total'] or 0

        balance = due.amount_due - paid
        if balance > 0:
            # Fetch correct academic record
            record_filter = {}
            if year_id:
                record_filter['academic_year_id'] = year_id
            if class_id_val:
                record_filter['class_enrolled_id'] = class_id_val

            academic_record = due.student.academic_records.filter(**record_filter).first()

            defaulters.append({
                'student': due.student,
                'class': academic_record.class_enrolled.name if academic_record else '',
                'month': due.month.strftime('%B %Y'),
                'fee_type': due.fee_type.name,
                'due': due.amount_due,
                'paid': paid,
                'balance': balance,
            })

    return render(request, 'reports/fees/fee_defaulter_report.html', {
        'dues': defaulters,
        'academic_years': academic_years,
        'classes': classes,
        'selected_year': academic_year,
        'selected_class': class_id,
        'selected_month': month_str,
    })



import openpyxl
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from fees.models import StudentFeeDue, StudentFeePaymentDetail
from django.db.models import Sum

@login_required
def export_fee_defaulter_report(request):
    user_school = request.user.userprofile.school
    academic_year = request.GET.get('academic_year')
    class_id = request.GET.get('class_id')
    month_str = request.GET.get('month')

    # Safe conversion
    year_id = int(academic_year) if academic_year and academic_year.isdigit() else None
    class_id_val = int(class_id) if class_id and class_id.isdigit() else None

    dues = StudentFeeDue.objects.select_related('student', 'fee_type').filter(student__school=user_school)

    if year_id:
        dues = dues.filter(student__academic_records__academic_year_id=year_id)
    if class_id_val:
        dues = dues.filter(student__academic_records__class_enrolled_id=class_id_val)
    if month_str:
        dues = dues.filter(month__month=int(month_str))

    rows = []
    for due in dues:
        paid = StudentFeePaymentDetail.objects.filter(
            payment__student=due.student,
            payment__month=due.month,
            fee_type=due.fee_type
        ).aggregate(total=Sum('amount_paid'))['total'] or 0

        balance = due.amount_due - paid
        if balance > 0:
            # Fetch academic record to get class
            record_filter = {}
            if year_id:
                record_filter['academic_year_id'] = year_id
            if class_id_val:
                record_filter['class_enrolled_id'] = class_id_val

            academic_record = due.student.academic_records.filter(**record_filter).first()
            class_name = academic_record.class_enrolled.name if academic_record else ''

            rows.append([
                due.student.full_name,
                class_name,
                due.month.strftime('%B %Y'),
                due.fee_type.name,
                float(due.amount_due),
                float(paid),
                float(balance)
            ])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fee Defaulters"

    ws.append(['Student', 'Class', 'Month', 'Fee Type', 'Due (₹)', 'Paid (₹)', 'Balance (₹)'])
    for row in rows:
        ws.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="fee_defaulters.xlsx"'
    wb.save(response)
    return response


# reports/views.py
from reports.forms import StudentFeeHistoryForm
from fees.models import StudentFeeDue, StudentFeePaymentDetail
from django.db.models import Sum
from collections import defaultdict

@login_required
def student_fee_history_report(request):
    school = request.user.userprofile.school
    form = StudentFeeHistoryForm(request.GET or None, school=school)
    history = []
    total_due = 0
    total_paid = 0

    if form.is_valid():
        student = form.cleaned_data['student']
        academic_year = form.cleaned_data['academic_year']

        dues = StudentFeeDue.objects.filter(student=student)
        if academic_year:
            dues = dues.filter(student__academic_records__academic_year=academic_year)

        payments = StudentFeePaymentDetail.objects.filter(payment__student=student)
        if academic_year:
            payments = payments.filter(payment__student__academic_records__academic_year=academic_year)

        # Group dues by (month, fee_type)
        due_dict = defaultdict(lambda: 0)
        for d in dues:
            key = (d.month, d.fee_type.name)
            due_dict[key] += float(d.amount_due)

        # Group payments by (month, fee_type)
        paid_dict = defaultdict(lambda: 0)
        for p in payments:
            key = (p.payment.month, p.fee_type.name)
            paid_dict[key] += float(p.amount_paid)

        all_keys = sorted(set(due_dict.keys()) | set(paid_dict.keys()))

        for key in all_keys:
            due = due_dict.get(key, 0)
            paid = paid_dict.get(key, 0)
            balance = due - paid
            total_due += due
            total_paid += paid

            history.append({
                'month': key[0].strftime('%B %Y'),
                'fee_type': key[1],
                'due': due,
                'paid': paid,
                'balance': balance
            })

    return render(request, 'reports/fees/student_fee_history.html', {
        'form': form,
        'history': history,
        'total_due': total_due,
        'total_paid': total_paid,
        'student': form.cleaned_data['student'] if form.is_valid() else None,
    })

# reports/views.py
import openpyxl
from django.http import HttpResponse
from collections import defaultdict
from fees.models import StudentFeeDue, StudentFeePaymentDetail

from django.contrib.auth.decorators import login_required

@login_required
def export_student_fee_history(request):
    school = request.user.userprofile.school
    student_id = request.GET.get('student')
    academic_year_id = request.GET.get('academic_year')

    try:
        student = StudentAdmission.objects.get(id=student_id, school=school)
    except StudentAdmission.DoesNotExist:
        return HttpResponse("Student not found.", status=404)

    academic_year = None
    if academic_year_id:
        try:
            academic_year = AcademicYear.objects.get(id=academic_year_id, school=school)
        except AcademicYear.DoesNotExist:
            pass  # Ignore invalid year

    dues = StudentFeeDue.objects.filter(student=student)
    if academic_year:
        dues = dues.filter(student__academic_records__academic_year=academic_year)

    payments = StudentFeePaymentDetail.objects.filter(payment__student=student)
    if academic_year:
        payments = payments.filter(payment__student__academic_records__academic_year=academic_year)

    # Group dues by (month, fee_type)
    due_dict = defaultdict(float)
    for d in dues:
        key = (d.month, d.fee_type.name)
        due_dict[key] += float(d.amount_due)

    # Group payments by (month, fee_type)
    paid_dict = defaultdict(float)
    for p in payments:
        key = (p.payment.month, p.fee_type.name)
        paid_dict[key] += float(p.amount_paid)

    all_keys = sorted(set(due_dict.keys()) | set(paid_dict.keys()))

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fee History"

    ws.append(['Student Name', student.full_name])
    ws.append(['Academic Year', academic_year.name if academic_year else 'All Years'])
    ws.append([])
    ws.append(['Month', 'Fee Type', 'Amount Due (₹)', 'Amount Paid (₹)', 'Balance (₹)'])

    total_due = total_paid = 0

    for key in all_keys:
        month, fee_type = key
        due = due_dict.get(key, 0)
        paid = paid_dict.get(key, 0)
        balance = due - paid
        total_due += due
        total_paid += paid

        ws.append([
            month.strftime('%B %Y'),
            fee_type,
            round(due, 2),
            round(paid, 2),
            round(balance, 2)
        ])

    # Totals
    ws.append([])
    ws.append(['Total Due', total_due])
    ws.append(['Total Paid', total_paid])
    ws.append(['Total Balance', total_due - total_paid])

    # Download response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"{student.full_name.replace(' ', '_')}_fee_history.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
